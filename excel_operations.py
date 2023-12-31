from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import range_boundaries
from openpyxl.utils import get_column_letter

def switch_columns(worksheet, col1_index, col2_index):
    """
    Switches the contents of two columns in an Openpyxl worksheet.

    Args:
    worksheet (Worksheet): The worksheet object from Openpyxl.
    col1_index (int): The index of the first column (1-based).
    col2_index (int): The index of the second column (1-based).
    """
    max_row = worksheet.max_row
    col1_letter, col2_letter = get_column_letter(col1_index), get_column_letter(col2_index)
    merged_ranges = []

    # Record merged cell ranges and unmerge them
    for merge_range in worksheet.merged_cells.ranges:
        if col1_letter in merge_range.coord or col2_letter in merge_range.coord:
            merged_ranges.append(str(merge_range))
            worksheet.unmerge_cells(str(merge_range))

    # Store column values and switch
    col1_values = [worksheet.cell(row, col1_index).value for row in range(1, max_row + 1)]
    col2_values = [worksheet.cell(row, col2_index).value for row in range(1, max_row + 1)]

    for row in range(1, max_row + 1):
        worksheet.cell(row, col1_index).value = col2_values[row - 1]
        worksheet.cell(row, col2_index).value = col1_values[row - 1]

    # Reapply merged cells with updated columns
    for merge_range in merged_ranges:
        new_range = merge_range.replace(col1_letter, 'temp').replace(col2_letter, col1_letter).replace('temp', col2_letter)
        worksheet.merge_cells(new_range)


def write_in_cells(worksheet, cell_range, text, color='000000', thickness='thin', font_name='Arial', font_size=12):
    """
    Writes text in a specified cell range, applies a border, and formats the text in an Openpyxl worksheet.

    Args:
    worksheet (Worksheet): The worksheet object from Openpyxl.
    cell_range (str): The cell range (e.g., 'A1:B10') to write the text in.
    text (str): The text to write in the cell range.
    color (str): The border color in hex format (e.g., '000000' for black).
    thickness (str): The border thickness (e.g., 'thin', 'thick').
    font_name (str): The font name to use.
    font_size (int): The font size to use.
    """
    # Merge cells
    worksheet.merge_cells(cell_range)

    # Define border style
    side = Side(border_style=thickness, color=color)
    border = Border(left=side, right=side, top=side, bottom=side)

    # Apply border to the entire range
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    for row in worksheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = border

    # Get the top-left cell of the merged area
    top_left_cell = worksheet.cell(row=min_row, column=min_col)

    # Set the text and alignment
    top_left_cell.value = text
    top_left_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Set the font
    top_left_cell.font = Font(name=font_name, size=font_size)

def set_outer_border(ws, cell_range=None, type='thin', color='000000'):
    """
    Apply a border to the perimeter of a specified cell range or the entire worksheet in an Openpyxl worksheet.

    Args:
    ws (openpyxl.worksheet.worksheet.Worksheet): The worksheet.
    cell_range (str, optional): The cell range (e.g., 'A1:B10'). If None, applies to the entire worksheet.
    type (str): The border style (e.g., 'thin', 'thick').
    color (str): The border color in hex format (e.g., 'FF0000' for red).
    """
    # Define the border style
    border_style = Side(border_style=type, color=color)

    if cell_range:
        # Apply the border to the specified cell range
        rows = list(ws[cell_range])
        min_row = rows[0][0].row
        max_row = rows[-1][0].row
        min_col = rows[0][0].column
        max_col = rows[0][-1].column
    else:
        # Apply the border to the entire worksheet
        min_row, max_row = 1, ws.max_row
        min_col, max_col = 1, ws.max_column

    # Apply top and bottom borders
    for col in range(min_col, max_col + 1):
        ws.cell(min_row, col).border = Border(top=border_style, left=ws.cell(min_row, col).border.left, right=ws.cell(min_row, col).border.right, bottom=ws.cell(min_row, col).border.bottom)
        ws.cell(max_row, col).border = Border(bottom=border_style, left=ws.cell(max_row, col).border.left, right=ws.cell(max_row, col).border.right, top=ws.cell(max_row, col).border.top)

    # Apply left and right borders
    for row in range(min_row, max_row + 1):
        ws.cell(row, min_col).border = Border(left=border_style, top=ws.cell(row, min_col).border.top, bottom=ws.cell(row, min_col).border.bottom, right=ws.cell(row, min_col).border.right)
        ws.cell(row, max_col).border = Border(right=border_style, top=ws.cell(row, max_col).border.top, bottom=ws.cell(row, max_col).border.bottom, left=ws.cell(row, max_col).border.left)

def remove_inner_borders(ws, cell_range):
    """
    Remove all borders inside a specified cell range in an Openpyxl worksheet.

    Args:
    ws (openpyxl.worksheet.worksheet.Worksheet): The worksheet.
    cell_range (str): The cell range (e.g., 'A1:B10').
    """
    no_border = Border(left=None, right=None, top=None, bottom=None)

    rows = list(ws[cell_range])
    min_row = rows[0][0].row
    max_row = rows[-1][0].row
    min_col = rows[0][0].column
    max_col = rows[0][-1].column

    # Iterate over cells within the range and remove borders
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            ws.cell(row, col).border = no_border

def fill_cells(ws, cell_range, type, color):
    """
    Fill a specified cell range in an Openpyxl worksheet with a color.

    Args:
    ws (openpyxl.worksheet.worksheet.Worksheet): The worksheet.
    cell_range (str): The cell range (e.g., 'A1:B10').
    type (str): The fill pattern type (e.g., 'solid').
    color (str): The fill color in hex format (e.g., 'FFEE00' for yellow).
    """
    # Define the fill pattern
    fill_pattern = PatternFill(fill_type=type, start_color=color, end_color=color)

    # Apply the fill to the specified cell range
    rows = ws[cell_range]
    for row in rows:
        for cell in row:
            cell.fill = fill_pattern


def auto_adjust_column_width(worksheet):
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column_letter].width = adjusted_width

def cell_alignment(worksheet, cell_range=None, Horizontal='center', Vertical='center'):
    """
    Apply alignment settings to a specified cell range or all non-empty cells in an Openpyxl worksheet.

    Args:
    worksheet (openpyxl.worksheet.worksheet.Worksheet): The worksheet.
    cell_range (str, optional): The cell range (e.g., 'A1:B10'). If None, applies to all non-empty cells.
    Horizontal (str): Horizontal alignment ('center', 'left', 'right', etc.).
    Vertical (str): Vertical alignment ('center', 'top', 'bottom', etc.).
    """

    # Define the alignment
    alignment_style = Alignment(horizontal=Horizontal, vertical=Vertical)

    if cell_range:
        # Apply the alignment to the specified cell range
        for row in worksheet[cell_range]:
            for cell in row:
                cell.alignment = alignment_style
    else:
        # Apply the alignment to all non-empty cells
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell.alignment = alignment_style