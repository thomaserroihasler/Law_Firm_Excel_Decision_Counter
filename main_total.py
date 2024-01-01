import pandas as pd
from collections import defaultdict
from collections import Counter

from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import range_boundaries
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList

import re
import os
import sys

def get_executable_path():
    return os.path.dirname(os.path.abspath(sys.argv[0]))


def extract_decisions_and_cases(title):
    """
    Extracts decision and case information from a given title string.

    This function searches for patterns that match legal decisions and cases,
    formatted as 'CAS/TAS YYYY XXX NNN', where 'CAS' or 'TAS' are decision
    prefixes, 'YYYY' is a year, 'XXX' is a series of letters, and 'NNN' is a number.
    The function handles cases where multiple decisions are concatenated with '+'.

    Args:
    title (str): The title string from which to extract decisions and cases.

    Returns:
    list: A list of extracted decisions and cases, formatted as strings.
    """

    # Regex to find all instances of decisions and cases within the title
    decisions = re.findall(r'\b(CAS|TAS)\s(\d{4}\s[A-Z]{1,3}\s\d{1,9}(?:\s?\+\s?\d{1,9})*)', title)
    all_cases = []

    for parts in decisions:
        prefix = parts[0]  # 'CAS' or 'TAS'

        # Normalize and split the case sequences
        case_sequence = re.sub(r'\s?\+\s?', ' + ', parts[1]).split(' + ')
        previous_year = 0000
        previous_letters = 'X'

        for case in case_sequence:
            part = case.split()

            # Handling full case format
            if len(part) == 3:
                year, letters, number = part
                previous_year = year
                previous_letters = letters
            # Handling case format with only number
            elif len(part) == 1:
                year, letters, number = previous_year, previous_letters, part[0]

            # Construct the full case string
            all_cases.append(f"{prefix} {year} {letters} {number}")

    return all_cases

def switch_columns(worksheet, col1_index, col2_index):
    """
    Switches the contents of two columns in an Openpyxl worksheet.

    Args:
    worksheet (Worksheet): The worksheet object from Openpyxl.
    col1_index (int): The index of the first column (1-based).
    col2_index (int): The index of the second column (1-based).
    """
    max_row = worksheet.max_row
    col1_letter, col2_letter = get_column_letter(col1_index), get_column_letter(col2_index)
    merged_ranges = []

    # Record merged cell ranges and unmerge them
    for merge_range in worksheet.merged_cells.ranges:
        if col1_letter in merge_range.coord or col2_letter in merge_range.coord:
            merged_ranges.append(str(merge_range))
            worksheet.unmerge_cells(str(merge_range))

    # Store column values and switch
    col1_values = [worksheet.cell(row, col1_index).value for row in range(1, max_row + 1)]
    col2_values = [worksheet.cell(row, col2_index).value for row in range(1, max_row + 1)]

    for row in range(1, max_row + 1):
        worksheet.cell(row, col1_index).value = col2_values[row - 1]
        worksheet.cell(row, col2_index).value = col1_values[row - 1]

    # Reapply merged cells with updated columns
    for merge_range in merged_ranges:
        new_range = merge_range.replace(col1_letter, 'temp').replace(col2_letter, col1_letter).replace('temp', col2_letter)
        worksheet.merge_cells(new_range)


def write_in_cells(worksheet, cell_range, text, color='000000', thickness='thin', font_name='Arial', font_size=12):
    """
    Writes text in a specified cell range, applies a border, and formats the text in an Openpyxl worksheet.

    Args:
    worksheet (Worksheet): The worksheet object from Openpyxl.
    cell_range (str): The cell range (e.g., 'A1:B10') to write the text in.
    text (str): The text to write in the cell range.
    color (str): The border color in hex format (e.g., '000000' for black).
    thickness (str): The border thickness (e.g., 'thin', 'thick').
    font_name (str): The font name to use.
    font_size (int): The font size to use.
    """
    # Merge cells
    worksheet.merge_cells(cell_range)

    # Define border style
    side = Side(border_style=thickness, color=color)
    border = Border(left=side, right=side, top=side, bottom=side)

    # Apply border to the entire range
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    for row in worksheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = border

    # Get the top-left cell of the merged area
    top_left_cell = worksheet.cell(row=min_row, column=min_col)

    # Set the text and alignment
    top_left_cell.value = text
    top_left_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Set the font
    top_left_cell.font = Font(name=font_name, size=font_size)

def set_outer_border(ws, cell_range=None, type='thin', color='000000'):
    """
    Apply a border to the perimeter of a specified cell range or the entire worksheet in an Openpyxl worksheet.

    Args:
    ws (openpyxl.worksheet.worksheet.Worksheet): The worksheet.
    cell_range (str, optional): The cell range (e.g., 'A1:B10'). If None, applies to the entire worksheet.
    type (str): The border style (e.g., 'thin', 'thick').
    color (str): The border color in hex format (e.g., 'FF0000' for red).
    """
    # Define the border style
    border_style = Side(border_style=type, color=color)

    if cell_range:
        # Apply the border to the specified cell range
        rows = list(ws[cell_range])
        min_row = rows[0][0].row
        max_row = rows[-1][0].row
        min_col = rows[0][0].column
        max_col = rows[0][-1].column
    else:
        # Apply the border to the entire worksheet
        min_row, max_row = 1, ws.max_row
        min_col, max_col = 1, ws.max_column

    # Apply top and bottom borders
    for col in range(min_col, max_col + 1):
        ws.cell(min_row, col).border = Border(top=border_style, left=ws.cell(min_row, col).border.left, right=ws.cell(min_row, col).border.right, bottom=ws.cell(min_row, col).border.bottom)
        ws.cell(max_row, col).border = Border(bottom=border_style, left=ws.cell(max_row, col).border.left, right=ws.cell(max_row, col).border.right, top=ws.cell(max_row, col).border.top)

    # Apply left and right borders
    for row in range(min_row, max_row + 1):
        ws.cell(row, min_col).border = Border(left=border_style, top=ws.cell(row, min_col).border.top, bottom=ws.cell(row, min_col).border.bottom, right=ws.cell(row, min_col).border.right)
        ws.cell(row, max_col).border = Border(right=border_style, top=ws.cell(row, max_col).border.top, bottom=ws.cell(row, max_col).border.bottom, left=ws.cell(row, max_col).border.left)

def remove_inner_borders(ws, cell_range):
    """
    Remove all borders inside a specified cell range in an Openpyxl worksheet.

    Args:
    ws (openpyxl.worksheet.worksheet.Worksheet): The worksheet.
    cell_range (str): The cell range (e.g., 'A1:B10').
    """
    no_border = Border(left=None, right=None, top=None, bottom=None)

    rows = list(ws[cell_range])
    min_row = rows[0][0].row
    max_row = rows[-1][0].row
    min_col = rows[0][0].column
    max_col = rows[0][-1].column

    # Iterate over cells within the range and remove borders
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            ws.cell(row, col).border = no_border

def fill_cells(ws, cell_range, type, color):
    """
    Fill a specified cell range in an Openpyxl worksheet with a color.

    Args:
    ws (openpyxl.worksheet.worksheet.Worksheet): The worksheet.
    cell_range (str): The cell range (e.g., 'A1:B10').
    type (str): The fill pattern type (e.g., 'solid').
    color (str): The fill color in hex format (e.g., 'FFEE00' for yellow).
    """
    # Define the fill pattern
    fill_pattern = PatternFill(fill_type=type, start_color=color, end_color=color)

    # Apply the fill to the specified cell range
    rows = ws[cell_range]
    for row in rows:
        for cell in row:
            cell.fill = fill_pattern


def auto_adjust_column_width(worksheet):
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column_letter].width = adjusted_width
from openpyxl.styles import Alignment

def cell_alignment(worksheet, cell_range=None, Horizontal='center', Vertical='center'):
    """
    Apply alignment settings to a specified cell range, an entire column, or all non-empty cells in an Openpyxl worksheet.

    Args:
    worksheet (openpyxl.worksheet.worksheet.Worksheet): The worksheet.
    cell_range (str, optional): The cell range (e.g., 'A1:B10') or column letter (e.g., 'A'). If None, applies to all non-empty cells.
    Horizontal (str): Horizontal alignment ('center', 'left', 'right', etc.).
    Vertical (str): Vertical alignment ('center', 'top', 'bottom', etc.).
    """

    # Define the alignment
    alignment_style = Alignment(horizontal=Horizontal, vertical=Vertical)

    if cell_range:
        # Check if cell_range is a single column letter
        if len(cell_range) == 1 and cell_range.isalpha():
            # Iterate through each row in the column
            for row in range(1, worksheet.max_row + 1):
                cell = worksheet[f'{cell_range}{row}']
                if cell.value is not None:  # Align only non-empty cells
                    cell.alignment = alignment_style
        else:
            # Apply the alignment to the specified cell range
            for row_cells in worksheet[cell_range]:
                for cell in row_cells:
                    cell.alignment = alignment_style
    else:
        # Apply the alignment to all non-empty cells
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell.alignment = alignment_style

def plot_frequency_bar_chart(workbook, worksheet_name, data, chart_title, chart_location='D1'):
    """
    Plots a frequency bar chart in an Excel worksheet using openpyxl and displays the frequency above each bar.

    Args:
        workbook (openpyxl.workbook.workbook.Workbook): An openpyxl Workbook object.
        worksheet_name (str): Name of the worksheet where the bar chart will be plotted.
        data (list): A list of data points for the bar chart.
        chart_title (str): Title of the bar chart.
        chart_location (str, optional): The cell location where the chart will be inserted. Defaults to 'D1'.
    """
    # Ensure the worksheet exists or create it
    if worksheet_name not in workbook.sheetnames:
        workbook.create_sheet(title=worksheet_name)
    worksheet = workbook[worksheet_name]
    data = sorted(data, key=lambda x: x[0], reverse=False)
    # Calculate starting row and column for data based on chart location
    chart_col = ord(chart_location[0]) - ord('A') + 1
    chart_row = int(chart_location[1:])
    start_row = chart_row
    start_col = chart_col  # Place data two columns to the left of the chart

    # Write the data to the worksheet
    for i, (category, frequency) in enumerate(data, start=1):
        worksheet.cell(row=start_row + i, column=start_col, value=category)
        worksheet.cell(row=start_row + i, column=start_col + 1, value=frequency)


    # Create a bar chart
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = chart_title
    chart.y_axis.title = 'Frequency'
    chart.x_axis.title = 'Year'

    # Add data to chart
    data_ref = Reference(worksheet, min_col=start_col + 1, min_row=start_row, max_row=start_row + len(data))
    cats_ref = Reference(worksheet, min_col=start_col, min_row=start_row + 1, max_row=start_row + len(data))
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)

    # Add data labels
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True  # Show the value on each bar

    # Place the chart on the worksheet at the specified location
    worksheet.add_chart(chart, chart_location)


def main():
    print("Current Working Directory:", get_executable_path())

    input_file_path = input("Please enter the name of the excel file (without extension): ")
    output_file_path = input("Please enter the name of the saved excel file (without extension): ")

    input_file_path += '.xlsx'
    output_file_path += '.xlsx'

    try:
        df = pd.read_excel(get_executable_path() + '/' + input_file_path)
    except Exception as e:
        print(f"An error occurred while reading the Excel file: {e}")
        return

    df = df.iloc[:, [0, 2]]  # Selecting specific columns
    df.iloc[:, 1] = pd.to_datetime(df.iloc[:, 1]).dt.date  # Convert to date-only format

    decision_case_counts = defaultdict(int)
    decisions = []

    for _, row in df.iterrows():
        decision_name = row.iloc[0]
        decision_date = row.iloc[1]
        cases = extract_decisions_and_cases(decision_name)
        cases_string = ' & '.join(cases)
        decisions.append([None, decision_name, decision_date, cases_string])
        decision_case_counts[(cases_string, decision_date)] += 1

    all_decisions_df = pd.DataFrame(decisions, columns=[None, 'Filename', 'Decision_Date', 'Cases_Numbers'])
    all_decisions_df = all_decisions_df.sort_values(by='Decision_Date', ascending=False)

    output_path = get_executable_path() + '/' + output_file_path
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        all_decisions_df.to_excel(writer, index=False, sheet_name='Cases_Numbers')

        workbook = writer.book
        worksheet = writer.sheets['Cases_Numbers']

        # Apply formatting, coloring, and border to decisions
        row_idx = 2  # Start from row 2 due to header
        counter = 1
        unique_years = []
        while row_idx <= len(all_decisions_df) + 1:
            row = all_decisions_df.iloc[row_idx - 2]
            decision_count = decision_case_counts[(row['Cases_Numbers'], row['Decision_Date'])]
            unique_years.append(row['Decision_Date'].year)
            end_row = row_idx + decision_count - 1

            cell_range = f'B{row_idx}:D{end_row}'
            cell_range_counter = f'A{row_idx}:A{end_row}'
            remove_inner_borders(worksheet, cell_range)
            if end_row != row_idx:
                fill_cells(worksheet, cell_range, 'solid', '00FF00')
                set_outer_border(worksheet, cell_range, 'thick')

            write_in_cells(worksheet, cell_range_counter, str(counter))
            set_outer_border(worksheet, cell_range_counter, 'thick')

            row_idx = end_row + 1
            counter += 1

        cell_alignment(worksheet,'A')
        cell_alignment(worksheet, 'C')
        cell_alignment(worksheet, 'D')

        auto_adjust_column_width(worksheet)
        set_outer_border(worksheet, f'B2:B{row_idx-1}', 'thin')
        set_outer_border(worksheet, f'C2:C{row_idx-1}', 'thin')
        set_outer_border(worksheet, None, 'thick')
        set_outer_border(worksheet, f'A1:D1', 'thick')
        year_counts = Counter(unique_years)
        data_for_chart = list(year_counts.items())
        plot_frequency_bar_chart(workbook, 'Cases_Numbers', data_for_chart, 'Decision Date Frequency', 'F1')

        # Define subsets
        subset_1 = all_decisions_df[all_decisions_df['Filename'].str.contains('CAS Web Archives')]
        subset_2 = all_decisions_df[all_decisions_df['Filename'].str.contains('CAS Bull')]
        subset_3 = all_decisions_df[~all_decisions_df['Filename'].isin(subset_1['Filename']) & ~all_decisions_df['Filename'].isin(subset_2['Filename'])]

        # Write subsets to different sheets
        subsets = [subset_1,subset_2,subset_3]
        sheet_names = ['CAS Web Archives','CAS Bull','Other']
        for i, subset in enumerate(subsets, start=1):
            sheet_name = sheet_names[i-1]
            workbook.create_sheet(sheet_name)
            subset.to_excel(writer, index=False, sheet_name=sheet_name)

            subset_worksheet = writer.sheets[sheet_name]
            # Apply formatting and add counter to the first column
            row_idx = 2  # Start from the second row
            for _, row in subset.iterrows():
                write_in_cells(subset_worksheet, f'A{row_idx}', str(row_idx - 1))
                row_idx += 1
            cell_alignment(subset_worksheet, 'A')
            cell_alignment(subset_worksheet, 'C')
            cell_alignment(subset_worksheet, 'D')
            auto_adjust_column_width(subset_worksheet)
            set_outer_border(subset_worksheet, f'B2:B{row_idx - 1}', 'thin')
            set_outer_border(subset_worksheet, f'C2:C{row_idx - 1}', 'thin')
            set_outer_border(subset_worksheet, f'A1:D1', 'thick')
            set_outer_border(subset_worksheet, None, 'thick')

            # Plot frequency bar chart of decision dates
            if 'Decision_Date' in subset:
                # Extract year from each datetime object
                decision_years = [date.year for date in subset['Decision_Date'] if pd.notnull(date)]
                year_counts = Counter(decision_years)
                data_for_chart = list(year_counts.items())
                plot_frequency_bar_chart(workbook, sheet_name, data_for_chart, 'Decision Date Frequency','F1')

if __name__ == "__main__":
    main()
