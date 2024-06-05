
import os
import sys
import pandas as pd
import re
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
import io
import pygame
import datetime
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def ask_names_of_files():
    """
    Asks the user for the names of an existing Excel file and a new output file.
    Ensures the input file exists and is an Excel file.
    Ensures the output file name is different from the input file name.
    """
    while True:
        try:
            # Ask for the input file name (without extension)
            input_file = input("Please enter the name of the existing Excel file (without .xlsx extension): ") + ".xlsx"
            
            # Check if the input file exists and is an Excel file
            if not os.path.isfile(input_file):
                raise FileNotFoundError("The specified file does not exist.")
            
            if not input_file.endswith('.xlsx'):
                raise ValueError("The file must be an Excel file with a .xlsx extension.")
            
            break
        except FileNotFoundError as fnf_error:
            print(fnf_error)
        except ValueError as val_error:
            print(val_error)
    
    while True:
        try:
            # Ask for the output file name (without extension)
            output_file = input("Please enter the name for the new output Excel file (without .xlsx extension): ") + ".xlsx"
            
            # Check if the output file name is different from the input file name (case-sensitive)
            if output_file == input_file:
                raise ValueError("The output file name must be different from the input file name.")
            
            break
        except ValueError as val_error:
            print(val_error)
    
    return input_file, output_file

def excel_sheet_to_data_frame(file_name, sheet_name=0):
    """
    Reads an Excel sheet into a pandas DataFrame.
    
    Parameters:
    - file_name: str, the name of the Excel file.
    - sheet_name: str or int, the name or index of the sheet to read (default is the first sheet).
    
    Returns:
    - DataFrame: the content of the Excel sheet as a pandas DataFrame.
    """
    try:
        # Read the Excel file into a DataFrame
        df = pd.read_excel(file_name, sheet_name=sheet_name)
        return df
    except FileNotFoundError:
        print("The specified file does not exist.")
    except ValueError as e:
        print(f"Error reading the Excel file: {e}")
    except Exception as e:
        print(f"An unexpected error occurred: {e}")
    return None

def separate_into_sub_categories(df):
    """
    Separates the DataFrame into subcategories based on specific keywords in the 'Filename' column.
    
    Parameters:
    - df: pandas DataFrame, the original DataFrame containing a 'Filename' column.
    
    Returns:
    - df_archive: pandas DataFrame, containing rows where 'Filename' includes 'Web archives' (case insensitive).
    - df_bull: pandas DataFrame, containing rows where 'Filename' includes 'Bull' (case insensitive).
    """
    # Ensure the 'Filename' column exists in the DataFrame
    if 'Filename' not in df.columns:
        raise KeyError("The DataFrame must contain a 'Filename' column.")
    
    # Create boolean masks for filtering
    archive_mask = df['Filename'].str.contains('Web archives', case=False, na=False)
    bull_mask = df['Filename'].str.contains('Bull', case=False, na=False)
    
    # Filter the DataFrame into subcategories
    df_archive = df[archive_mask].copy()
    df_bull = df[bull_mask].copy()
    
    return df_archive, df_bull

def df_to_excel_sheet(df, file_name, sheet_name, mode='a'):
    """
    Writes a DataFrame into a new Excel worksheet in a file.
    
    Parameters:
    - df: pandas DataFrame, the DataFrame to write to the Excel file.
    - file_name: str, the name of the Excel file to write to.
    - sheet_name: str, the name of the sheet to write the DataFrame into.
    - mode: str, the mode to open the Excel file ('w' for write, 'a' for append).
    
    Returns:
    - None
    """
    try:
        # Use ExcelWriter with openpyxl engine to append to an existing Excel file or create a new one
        with pd.ExcelWriter(file_name, engine='openpyxl', mode=mode) as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
        print(f"DataFrame successfully written to {file_name} in sheet {sheet_name}.")
    except FileNotFoundError:
        print("The specified file does not exist.")
    except Exception as e:
        print(f"An unexpected error occurred: {e}")


def order_data_frame(df):
    """
    Orders the DataFrame by 'Extracted dates' in ascending order.
    Additional ordering criteria or processing can be added as needed.
    
    Parameters:
    - df: pandas DataFrame, the DataFrame to be ordered.
    
    Returns:
    - df: pandas DataFrame, the ordered DataFrame.
    """
    # Check if the 'Extracted dates' column exists
    if 'Extracted dates' in df.columns:
        # Order the DataFrame by 'Extracted dates' in ascending order
        df = df.sort_values(by='Extracted dates', ascending=True)
    
    # Additional processing or ordering criteria can be added here
    
    return df


def extract_and_write_case_number(df):
    """
    Extracts the case number from the 'Filename' column and adds it as a new column 'Case Number'.
    If no match is found with 'CAS' or 'TAS', it tries to find the same pattern without these prefixes.
    
    Parameters:
    - df: pandas DataFrame, the original DataFrame containing filenames.
    
    Returns:
    - df: pandas DataFrame, the modified DataFrame with a new 'Case Number' column.
    """
    pattern_with_prefix = r'\(\s*(CAS|TAS) (\d{2,4})? ?([A-Z ]+)(.*?)\)'
    pattern_without_prefix = r'\(\s*(\d{2,4})? ?([A-Z ]+)(.*?)\)'

    # Define a function to extract the case number using the provided pattern
    def extract_case_number(s):
        match = re.search(pattern_with_prefix, s)
        if match:
            return match.group(0).strip('()')
        else:
            match = re.search(pattern_without_prefix, s)
            if match:
                return match.group(0).strip('()')
        return None

    # Apply the extraction function to the 'Filename' column
    df['Case Number'] = df.iloc[:, 0].apply(extract_case_number)

    # Check if any case numbers were extracted; raise an exception if not
    if df['Case Number'].isnull().all():
        raise ValueError("No case numbers matching the pattern were found.")

    return df

def proper_ordering_cases_by_case_number(df):
    """
    Orders the DataFrame by 'Extracted dates' and groups cases with the same last 6 symbols
    of 'Case Number' next to each other for each year.
    
    Parameters:
    - df: pandas DataFrame, the DataFrame to be ordered.
    
    Returns:
    - df: pandas DataFrame, the ordered DataFrame.
    """
    # Ensure 'Extracted dates' and 'Case Number' columns exist
    if 'Extracted dates' not in df.columns or 'Case Number' not in df.columns:
        raise KeyError("The DataFrame must contain 'Extracted dates' and 'Case Number' columns.")
    
    # Extract the year from 'Extracted dates'
    df['Year'] = df['Extracted dates'].apply(lambda x: x.year if pd.notnull(x) else None)
    
    # Extract the last 6 symbols of the 'Case Number' for grouping
    df['Short Case Number'] = df['Case Number'].apply(lambda x: str(x)[-6:] if pd.notna(x) else None)
    
    # Sort by 'Year' first to group by year, and then by 'Short Case Number' to group cases
    df = df.sort_values(by=['Year', 'Short Case Number'])
    
    # Drop the temporary 'Year' and 'Short Case Number' columns
    df.drop(columns=['Year', 'Short Case Number'], inplace=True)
    
    return df

def group_identical_decisions(df):
    """
    Groups rows with the same last 6 symbols of 'Case Number' and creates a counter next to each decision.
    The counter increments with each new unique set of last 6 symbols of 'Case Number' and remains the same if the case number repeats.
    
    Parameters:
    - df: pandas DataFrame, the DataFrame containing a 'Case Number' column.
    
    Returns:
    - df: pandas DataFrame, the modified DataFrame with a 'Decision Counter' column as the first column.
    """
    if 'Case Number' not in df.columns:
        raise KeyError("The DataFrame must contain a 'Case Number' column.")
    
    # Initialize the counter
    decision_counter = []
    current_case = None
    counter = 0

    # Iterate over the DataFrame rows
    for i in range(len(df)):
        case_number = df.at[i, 'Case Number']
        # Ensure the case number is a string and has at least 6 characters
        if pd.notna(case_number) and isinstance(case_number, str) and len(case_number) >= 6:
            last_six = case_number[-6:]
        else:
            last_six = None
        
        if last_six != current_case:
            current_case = last_six
            counter += 1
        decision_counter.append(counter)
    
    # Add the 'Decision Counter' column at the beginning
    df.insert(0, 'Decision Counter', decision_counter)
    
    return df

def stylize_file():
    return 0

def extract_dates_from_file_name_in_data_frame(df):
    """
    Extracts dates from the beginning of strings in the first column of the DataFrame 
    and creates a new column called 'Extracted dates' containing only the date.
    
    Parameters:
    - df: pandas DataFrame, the original DataFrame containing dates in the first column.
    
    Returns:
    - df: pandas DataFrame, the modified DataFrame with a new 'Extracted dates' column.
    """
    # Define a function to extract the date from the beginning of the string
    def extract_date(s):
        match = re.match(r'^(\d{4} \d{2} \d{2})', s)
        return pd.to_datetime(match.group(0), format='%Y %m %d').date() if match else pd.NaT
    
    # Apply the extraction function to the first column
    df['Extracted dates'] = df.iloc[:, 0].apply(extract_date)
    return df

def make_graph(df, output_file, sheetname='MainData', location='H2'):
    """
    Creates a bar plot showing the number of unique case numbers per year,
    with counts for each bar and the total amount on the plot. The plot is
    embedded in the Excel file.
    
    Parameters:
    - df: pandas DataFrame, the DataFrame containing the data.
    - output_file: str, the name of the output Excel file.
    - sheetname: str, the name of the sheet where the plot will be embedded.
    - location: str, the cell location where the plot will be embedded.
    
    Returns:
    - None
    """
    # Ensure 'Case Number' and 'Extracted dates' columns exist
    if 'Case Number' not in df.columns or 'Extracted dates' not in df.columns:
        raise KeyError("The DataFrame must contain 'Case Number' and 'Extracted dates' columns.")
    
    # Extract the year from 'Extracted dates'
    df['Year'] = df['Extracted dates'].apply(lambda x: x.year if pd.notnull(x) else None)
    
    # Extract the last 6 symbols of the 'Case Number' for analysis
    df['Short Case Number'] = df['Case Number'].apply(lambda x: str(x)[-6:] if pd.notna(x) else None)
    
    # Drop duplicates to keep only unique case numbers per year
    df_unique_cases = df.drop_duplicates(subset=['Short Case Number', 'Year'])
    
    # Group by 'Year' and count unique case numbers
    yearly_case_counts = df_unique_cases.groupby('Year')['Short Case Number'].count()
    
    # Create a bar plot
    plt.figure(figsize=(10, 6))
    bars = yearly_case_counts.plot(kind='bar')
    
    # Add counts above each bar
    for bar in bars.patches:
        height = bar.get_height()
        bars.annotate(f'{height}', 
                      xy=(bar.get_x() + bar.get_width() / 2, height),
                      xytext=(0, 3),  # 3 points vertical offset
                      textcoords="offset points",
                      ha='center', va='bottom')
    
    # Calculate total amount
    total_amount = yearly_case_counts.sum()
    
    # Add total amount on the plot
    plt.title(f'Number of Unique Case Numbers per Year (Total: {total_amount})')
    plt.xlabel('Year')
    plt.ylabel('Number of Unique Case Numbers')
    plt.tight_layout()
    
    # Save the plot to a BytesIO object
    image_stream = io.BytesIO()
    plt.savefig(image_stream, format='png')
    plt.close()
    image_stream.seek(0)
    
    # Load the workbook and select the sheet
    workbook = load_workbook(output_file)
    sheet = workbook[sheetname]
    
    # Add the image to the sheet
    img = ExcelImage(image_stream)
    img.anchor = location  # Position the image at the specified location
    sheet.add_image(img)
    
    # Save the workbook
    workbook.save(output_file)
    
    print(f"Graph embedded in {output_file} at {location}")


from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

import datetime
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

base_path = os.path.dirname(os.path.abspath(sys.executable if getattr(sys, 'frozen', False) else __file__))

def make_worksheet_more_aesthetic(file_name, sheet_name):
    """
    Makes the specified worksheet more aesthetic by adjusting column widths,
    applying a bold font to the header row, centering the header text, centering the entire first column,
    highlighting rows with the same counter value in the first column, making the lines separating columns bold
    and rows normal, and removing the time part from column E and rewriting the date as a string in YYYY-MM-DD format.
    
    Parameters:
    - file_name: str, the name of the Excel file.
    - sheet_name: str, the name of the worksheet to format.
    
    Returns:
    - None
    """
    # Load the workbook and select the sheet
    workbook = load_workbook(file_name)
    sheet = workbook[sheet_name]
    
    # Set column widths
    for column in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)  # Get the column name
        
        # Iterate through all rows in the column to find the max length
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        # Set the column width
        adjusted_width = max_length  # Set buffer to 0
        sheet.column_dimensions[column_letter].width = adjusted_width
    
    # Apply a bold font and center alignment to the header row
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    for cell in sheet[1]:  # Assuming the first row is the header
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Center align the entire first column
    first_column_letter = get_column_letter(1)
    for cell in sheet[first_column_letter]:
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Remove time part from column E and rewrite date as a string in YYYY-MM-DD format
    column_e_letter = 'E'
    
    for cell in sheet[column_e_letter]:
        if isinstance(cell.value, datetime.datetime):
            cell.value = cell.value.strftime('%Y-%m-%d')
    
    # Find and highlight rows with the same counter value in the first column
    counter_color_map = {}
    current_color_index = 0
    colors = ['FFFF99', 'FFCC99', 'FF9999', '99CCFF', '99FF99']  # List of colors to cycle through
    
    for row in sheet.iter_rows(min_row=2, max_col=sheet.max_column, max_row=sheet.max_row):  # Skip header row
        counter_value = row[0].value
        if counter_value in counter_color_map:
            fill = counter_color_map[counter_value]
        else:
            fill = PatternFill(start_color=colors[current_color_index % len(colors)], end_color=colors[current_color_index % len(colors)], fill_type='solid')
            counter_color_map[counter_value] = fill
            current_color_index += 1
        
        for cell in row:
            cell.fill = fill

    # Define the bold border style for columns and normal border style for rows
    bold_side = Side(style='medium')
    normal_side = Side(style='thin')
    
    for row in sheet.iter_rows():
        for cell in row:
            cell.border = Border(
                left=bold_side if cell.column == 1 else normal_side,
                right=bold_side if cell.column == sheet.max_column else normal_side,
                top=normal_side,
                bottom=normal_side
            )
    
    # Save the workbook with changes
    workbook.save(file_name)
    print(f"Worksheet '{sheet_name}' in '{file_name}' has been formatted.")

def ask_names_of_files_ui():
    """
    Uses Pygame to create a UI for the user to input the names of the input and output Excel files.
    """
    pygame.init()
    screen = pygame.display.set_mode((600, 300))
    pygame.display.set_caption("Excel File Processor")

    font = pygame.font.Font(None, 32)
    input_box = pygame.Rect(150, 50, 350, 32)
    output_box = pygame.Rect(150, 100, 350, 32)
    confirm_button = pygame.Rect(250, 200, 100, 40)
    color_inactive = pygame.Color('lightskyblue3')
    color_active = pygame.Color('dodgerblue2')
    color_error = pygame.Color('red')
    button_color = pygame.Color('green')
    color = color_inactive
    active = 'input'
    text_input = ''
    text_output = ''
    done = False
    input_valid = True
    output_valid = True

    while not done:
        for event in pygame.event.get():
            
            if event.type == pygame.QUIT:
                done = True
                pygame.quit()
                return None, None
            
            if event.type == pygame.MOUSEBUTTONDOWN:
                if input_box.collidepoint(event.pos):
                    active = 'input'
                elif output_box.collidepoint(event.pos):
                    active = 'output'
                elif confirm_button.collidepoint(event.pos):
                    input_file = text_input + ".xlsx"
                    output_file = text_output + ".xlsx"
                    input_valid = os.path.isfile(input_file)
                    output_valid = not os.path.isfile(output_file)
                    if input_valid and output_valid:
                        done = True
                else:
                    active = None
                color = color_active if active else color_inactive
            if event.type == pygame.KEYDOWN:
                if active == 'input':
                    input_valid = True  # Reset validity on new input
                    if event.key in [pygame.K_RETURN, pygame.K_TAB]:
                        active = 'output'
                    elif event.key == pygame.K_BACKSPACE:
                        text_input = text_input[:-1]
                    else:
                        text_input += event.unicode
                elif active == 'output':
                    output_valid = True  # Reset validity on new input
                    if event.key == pygame.K_RETURN:
                        input_file = base_path+'/'+text_input + ".xlsx"
                        output_file = base_path+'/'+text_output + ".xlsx"
                        print(f"Current working directory: {os.getcwd()}")
                        print(f"Input file path: {os.path.abspath(input_file)}")
                        print(f"Output file path: {os.path.abspath(output_file)}")
                        
                        print(base_path)
                        input_valid = os.path.isfile(input_file)
                        output_valid = not os.path.isfile(output_file)
                        if input_valid and output_valid:
                            done = True
                    elif event.key == pygame.K_TAB:
                        active = 'input'
                    elif event.key == pygame.K_BACKSPACE:
                        text_output = text_output[:-1]
                    else:
                        text_output += event.unicode

        screen.fill((30, 30, 30))
        txt_surface_input = font.render(text_input, True, color_active if active == 'input' else color_inactive)
        txt_surface_output = font.render(text_output, True, color_active if active == 'output' else color_inactive)
        screen.blit(txt_surface_input, (input_box.x + 5, input_box.y + 5))
        screen.blit(txt_surface_output, (output_box.x + 5, output_box.y + 5))
        
        # Draw input box with appropriate border color
        #input_box_color = color_active if active == 'input' else (color_error if not input_valid else color_inactive)
        #output_box_color = color_active if active == 'output' else (color_error if not output_valid else color_inactive)
        input_box_color = color_error if not input_valid else color_inactive
        output_box_color = color_error if not output_valid else color_inactive
        #(output_valid,active)
        pygame.draw.rect(screen, input_box_color, input_box, 2)
        pygame.draw.rect(screen, output_box_color, output_box, 2)

        # Draw confirm button
        pygame.draw.rect(screen, button_color, confirm_button)
        confirm_text = font.render('Confirm', True, (255, 255, 255))
        screen.blit(confirm_text, (confirm_button.x + 10, confirm_button.y + 10))

        # Render instructions
        instructions = font.render('Enter input and output file names (without .xlsx):', True, pygame.Color('white'))
        screen.blit(instructions, (10, 10))

        # Render labels
        input_label = font.render('Input:', True, pygame.Color('white'))
        output_label = font.render('Output:', True, pygame.Color('white'))
        screen.blit(input_label, (50, 50))
        screen.blit(output_label, (50, 100))

        pygame.display.flip()

    pygame.quit()
    return input_file, output_file


def main():
    # Get input and output file names using Pygame UI
    input_file, output_file = ask_names_of_files_ui()
    
    if input_file is None or output_file is None:
        return
    
    # # Print the results
    # print(f"Input file: {input_file}")
    # print(f"Output file: {output_file}")
    
    # Create the output file
    with pd.ExcelWriter(output_file, mode='w') as writer:
        pass

    # Convert the first sheet of the input file into a DataFrame
    df = excel_sheet_to_data_frame(input_file)
    df_archive = pd.DataFrame()  # empty data frame
    df_bull = pd.DataFrame()  # empty data frame
    if df is not None:
        # Rename the third column as "Kettle decision dates"
        df.rename(columns={df.columns[2]: 'Kettle decision dates'}, inplace=True)
        df = extract_dates_from_file_name_in_data_frame(df)
        df = extract_and_write_case_number(df)

        df = order_data_frame(df)
        df = proper_ordering_cases_by_case_number(df)
        print("Data from the first sheet:")
        #print(df)
        
        # Create the output file and write the main DataFrame as the first sheet
        df_to_excel_sheet(df, output_file, 'MainData', mode='w')
        # Separate into subcategories
        df_archive, df_bull = separate_into_sub_categories(df)
       #print(df_archive, df_bull)
        df_to_excel_sheet(df_archive, output_file, 'Archive', mode='a')
        df_to_excel_sheet(df_bull, output_file, 'Bull', mode='a')
        # Write DataFrames to the output file
       
    df = excel_sheet_to_data_frame(output_file)
    df_archive = excel_sheet_to_data_frame(output_file, 'Archive')
    df_bull = excel_sheet_to_data_frame(output_file, 'Bull')
    
    if df is not None:
        df = group_identical_decisions(df)
        df_archive = group_identical_decisions(df_archive)
        df_bull = group_identical_decisions(df_bull)
        df_to_excel_sheet(df, output_file, 'MainData', mode='w')
        df_to_excel_sheet(df_archive, output_file, 'Archive', mode='a')
        df_to_excel_sheet(df_bull, output_file, 'Bull', mode='a')
        make_graph(df, output_file, 'MainData')
        make_graph(df_archive, output_file, 'Archive')
        make_graph(df_bull, output_file, 'Bull')
        make_worksheet_more_aesthetic(output_file, 'MainData')
        make_worksheet_more_aesthetic(output_file, 'Archive')
        make_worksheet_more_aesthetic(output_file, 'Bull')
    # Print the final DataFrame
    # print("Final Data from the first sheet:")
    # print(df)

if __name__ == "__main__":
    main()

# def main():
#     # Get input and output file names
#     input_file, output_file = ask_names_of_files()
    
#     # Print the results
#     print(f"Input file: {input_file}")
#     print(f"Output file: {output_file}")
    
#     # Create the output file
#     with pd.ExcelWriter(output_file, mode='w') as writer:
#         pass

#     # Convert the first sheet of the input file into a DataFrame
#     df = excel_sheet_to_data_frame(input_file)
#     df_archive = pd.DataFrame()  # empty data frame
#     df_bull = pd.DataFrame()  # empty data frame
#     if df is not None:
#         # Rename the third column as "Kettle decision dates"
#         df.rename(columns={df.columns[2]: 'Kettle decision dates'}, inplace=True)
#         df = extract_dates_from_file_name_in_data_frame(df)
#         df = extract_and_write_case_number(df)

#         df = order_data_frame(df)
#         df = proper_ordering_cases_by_case_number(df)
#         print("Data from the first sheet:")
#         print(df)
        
#         # Create the output file and write the main DataFrame as the first sheet
#         df_to_excel_sheet(df, output_file, 'MainData', mode='w')
#         # Separate into subcategories
#         df_archive, df_bull = separate_into_sub_categories(df)
#         print(df_archive, df_bull)
#         df_to_excel_sheet(df_archive, output_file, 'Archive', mode='a')
#         df_to_excel_sheet(df_bull, output_file, 'Bull', mode='a')
#         # Write DataFrames to the output file
       
#     df =  excel_sheet_to_data_frame(output_file)
#     df_archive =  excel_sheet_to_data_frame(output_file,'Archive')
#     df_bull =  excel_sheet_to_data_frame(output_file,'Bull')
    
#     if df is not None:
#         df = group_identical_decisions(df)
#         df_archive = group_identical_decisions(df_archive)
#         df_bull = group_identical_decisions(df_bull)
#         df_to_excel_sheet(df, output_file, 'MainData', mode='w')
#         df_to_excel_sheet(df_archive, output_file, 'Archive', mode='a')
#         df_to_excel_sheet(df_bull, output_file, 'Bull', mode='a')
#         make_graph(df,output_file,'MainData')
#         make_graph(df_archive,output_file,'Archive')
#         make_graph(df_bull,output_file,'Bull')
#         make_worksheet_more_aesthetic(output_file, 'MainData')
#         make_worksheet_more_aesthetic(output_file, 'Archive')
#         make_worksheet_more_aesthetic(output_file, 'Bull')
#     #     print(df)
# if __name__ == "__main__":
#     main()

